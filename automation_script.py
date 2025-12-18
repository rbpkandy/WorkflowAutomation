import pandas as pd
import logging
from pathlib import Path
from openpyxl import load_workbook, chart #Import load_workbook from openpyxl
from datetime import datetime

#Root directory based on the script's location
ROOT_DIR = Path(__file__).parent
MASTER_DATA_DIR = ROOT_DIR / "master_data"
EMPLOYEE_FILES_DIR = ROOT_DIR / "employee_files"
LOG_FILE = ROOT_DIR / f"automation_log_{datetime.now().strftime('%Y%m%d')}.log"

#logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)


def load_and_clean_master_data(file_path: Path) -> pd.DataFrame or None:
    """Loads the master dataset and performs cleaning/validation."""
    logging.info(f"Attempting to load master data from: {file_path}")
    try:
        #Assume master data is in the first sheet
        df = pd.read_excel(file_path)
        
        #Data Cleaning/Transformation steps
        df.columns = df.columns.str.strip().str.replace(' ', '_')
        df = df.rename(columns={'Employee_ID': 'EmpID'}) 
        df['EmpID'] = df['EmpID'].astype(str).str.upper().str.strip()
        
        #Filter required columns
        required_cols = ['EmpID', 'Date', 'Pay_Component', 'Amount']
        df = df[required_cols]

        logging.info(f"Master data loaded successfully. Total rows: {len(df)}")
        return df

    except FileNotFoundError:
        logging.error(f"FATAL ERROR: Master data file not found at {file_path}")
        return None
    except Exception as e:
        logging.error(f"An error occurred during master data processing: {e}")
        return None


def append_to_employee_file(emp_df: pd.DataFrame, emp_id: str, target_sheet: str = 'Raw_Data'):
    """Append new data to an employee-specific Excel file, preserving structure."""
    
    file_name = f"EMP_{emp_id}_Report.xlsx"
    target_path = EMPLOYEE_FILES_DIR / file_name
    
    logging.info(f"Processing data for Employee ID: {emp_id}")

    if not target_path.exists():
        logging.warning(f"Target file not found for {emp_id}. Skipping append.")
        return

    try:
        #To preserve formulas, charts, and formatting, load the existing workbook structure *without* loading data immediately
        book = load_workbook(target_path)
        
        if target_sheet not in book.sheetnames:
            logging.warning(f"Sheet '{target_sheet}' not found in {file_name}. Skipping append.")
            return

        writer = pd.ExcelWriter(target_path, engine='openpyxl')
        writer.book = book # Assign the loaded workbook to the writer
        
        #Read existing data from the target sheet to determine where to start writing
        #Use pandas to read the existing data
    
        try:
             existing_df = pd.read_excel(target_path, sheet_name=target_sheet)
             #Starting row for the new data (index is 0-based, plus 1 for header)
             start_row = len(existing_df) + 1 
        except Exception:
             #Handle case where sheet is empty
             start_row = 1 


        #Append the new data
        emp_df.to_excel(
            writer,
            sheet_name=target_sheet,
            index=False,
            header=False,        #Don't write the header again
            startrow=start_row   #Start writing immediately after the existing data
        )
        
        writer.close()
        logging.info(f"Successfully appended {len(emp_df)} rows to {file_name} starting at row {start_row + 1}.") 

    except Exception as e:
        logging.error(f"ERROR appending data to {file_name}: {e}")
        


def main():
    """Manages the overall ETL process."""
    logging.info("--- Starting Data Automation Process ---")
    
    #ETL    
    master_files = list(MASTER_DATA_DIR.glob('*.xlsx'))
    if not master_files:
        logging.error("No master data Excel files found in the master_data directory. Exiting.")
        return

    master_file_path = master_files[0]
    master_df = load_and_clean_master_data(master_file_path)

    if master_df is None:
        logging.error("Master data could not be processed. Exiting.")
        return

    #Grouping and Loading
    grouped_data = master_df.groupby('EmpID')
    
    total_files_processed = 0
    
    #Iterate through each employee's data
    for emp_id, data_group in grouped_data:
        append_to_employee_file(data_group, emp_id)
        total_files_processed += 1

    logging.info(f"--- Process Complete ---")
    logging.info(f"Total employee files processed: {total_files_processed}")


if __name__ == '__main__':
    #Ensure directories exist before running
    MASTER_DATA_DIR.mkdir(exist_ok=True)
    EMPLOYEE_FILES_DIR.mkdir(exist_ok=True)
    
    #Print instruction
    print("\nStarting Excel Automation. Check the log file for details after completion.")
    
    main()

